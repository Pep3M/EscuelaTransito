from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from os import mkdir, path

class CursoModelo():

    def __init__(self, matricula_inicial) -> None:
        TEMPLATE_CURSO_URL = 'assets/templates/Curso MODELO.xlsx'
        self.wb = load_workbook(TEMPLATE_CURSO_URL)
        self.ws = self.wb.active
        self.ws['A9'].value = matricula_inicial


    def fecha_horario(self, fecha_inicio, fecha_final, horario):
        self.ws['A3'].value = f'Fecha inicio: {fecha_inicio}                                                       Fecha Termino: {fecha_final}                                Grupo: cp'
        self.ws['A4'].value = f'Instructor:     Yamilet Reyes Valdes                                                   Firma:                                           Horario:   {horario}'


    def agregar_matricula(self, nombre:str, dni:str=''):
        # buscamos una celda vacia para agregar la matricula
        for i in range(9,44):
            
            if self.ws[f'B{i}'].value == None:
                # agregamos el nombre
                self.ws[f'B{i}'].value = nombre
            
                # agregamos dni
                if not dni == '':    
                    for col in range(3,13):
                        try: self.ws[f'{get_column_letter(col)}{i}'].value = dni[col-3]
                        except: pass
                    try: self.ws[f'S{i}'].value = dni[10]
                    except:pass
                
                break


    def agregar_lote_matriculas(self, lista:list[list]):
        for item in lista:
            if type(item) == str:
                self.agregar_matricula(item)
            else:
                self.agregar_matricula(item[0],item[1])
            
    
    def exportar(self, file):
        self.wb.save(file)


class AulaModelo():

    def __init__(self, matricula_inicial) -> None:
        TEMPLATE_AULA_URL = 'assets/templates/Teorico para aula MODELO.xlsx'
        self.wb = load_workbook(TEMPLATE_AULA_URL)
        self.ws = self.wb.active
        self.ws['A9'].value = matricula_inicial


    def fecha_horario(self, fecha_inicio, fecha_final, horario):
        self.ws['A3'].value = f'Fecha inicio: {fecha_inicio}                                                       Fecha Termino: {fecha_final}                                Grupo: cp'
        self.ws['A4'].value = f'Instructor:     Yamilet Reyes Valdes                                                   Firma:                                           Horario:   {horario}'


    def agregar_matricula(self, nombre:str, dni:str='', tel='', unidad=''):
        # buscamos una celda vacia para agregar la matricula
        for i in range(9,44):
            
            if self.ws[f'B{i}'].value == None:
                # agregamos el nombre
                self.ws[f'B{i}'].value = nombre
            
                # agregamos dni
                if not dni == '':    
                    for col in range(3,14):
                        try: self.ws[f'{get_column_letter(col)}{i}'].value = dni[col-3]
                        except: pass
                
                #tel
                if not tel == '':
                    self.ws[f'N{i}'].value = tel
                
                #unidad
                if not unidad == '':
                    self.ws[f'P{i}'].value = unidad
                
                break


    def agregar_lote_matriculas(self, lista:list[list]):
        for item in lista:
            if type(item) == str:
                self.agregar_matricula(item)
            else:
                self.agregar_matricula(item[0],item[1],item[2],item[3])
            
    
    def exportar(self, file):
        self.wb.save(file)





def crear_curso_prueba():
    # inicializamos el curso
    curso1 = CursoModelo(134)
    curso1.fecha_horario(fecha_inicio='1/06/2022', fecha_final='10/07/2022', horario='11:00am - 1:00pm')

    # pasamos un lote de matriculas
    lista = [
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
        ['juan lopez AGUIRRE', '91011627921'],
        ['Alberto Gomez', '81061327921'],
        ['Luis Abreu', '97123127921'],
    ]
    curso1.agregar_lote_matriculas(lista)

    # guardamos el excel en un archivo
    if not path.exists('export'): mkdir('export')
    curso1.exportar('export/curso.xlsx')